#!/usr/bin/env python3
"""
유형 예측표 생성기 — 45단원 분류표 → predicted_axes   [스크립트 v5]

사용:  python build_axis_prediction.py <분류표.xlsx> <PREFIX> [출력.xlsx]
예:    python build_axis_prediction.py 벡터_수정본_v2.xlsx GEV
       python build_axis_prediction.py 확률_수정본_v2.xlsx PSP 예측표_PSP.xlsx

원칙(인계문 §16):
  - 값을 지어내지 않는다. 규칙에 걸린 것만 출력한다.
  - 걸리지 않으면 공란으로 두고 unmatched 로 보고한다.
  - 출력은 predicted_ 전용. 학생 실측(observed_)과 절대 섞지 않는다.
  - 규칙이 바뀌면 전량 재생성한다. 수동 편집 금지.
  - 원본(1층)은 읽기만 한다. 스크립트가 분류표를 고치는 일은 없다.

v5 변경:
  0. 매칭 컨텍스트에서 중영역을 제거한다. 이제 「유형 묶음 + 유형 이름」만 본다.
     45단원 실측: 공통규칙 발화 12,820건 중 이름에서 걸린 것은 7,497건(58%)뿐이고
     1,813건은 중영역만으로 걸렸다. 중영역은 단원의 주제어라서 EQ「일차방정식」·
     IN「일차부등식」·CMM「행렬의 연산」처럼 전 항목에 같은 축을 붙여 적중률만
     부풀린다. 제거 비용은 pack_gap 6,410행 중 373행(6%)이 미매칭으로 내려가는
     것이고, 표본 점검 결과 대부분 축이 틀린 오탐이었다.
     중영역은 unmatched·pack_gap 시트에 계속 실어 규칙 설계 재료로만 쓴다.

v4 변경:
  0. 유형요약 시트도 읽는다. 그룹ID 가 문항분류표에 없는 묶음(§16 배지없음)을
     유형묶음 출처 1행으로 방출한다. 이게 없으면 정본 산출물이 엔진 데이터보다
     커버리지가 낮아진다(45단원 156건). 이 행들은 주제코드가 없어 그룹ID 로 조인.
  1. pack_gap 시트 추가. 「걸렸지만 공통규칙만 걸린 항목」 — 팩 필요성의 실제 지표.
     CMM 은 미매칭 2건(3%)인데 공통규칙만 59건(97%)이었다. unmatched 만 보면
     과다매칭이 안 보인다. 규칙 보강의 재료는 unmatched 가 아니라 이 시트다.
  2. meta 에 팩 기여 / 공통규칙만 / 미매칭 3분할 추가.

v3 변경:
  0. detect_layout → resolve_name. 단원마다 층을 하나 고르는 방식을 폐기하고
     행별로 세부유형 → 주제유형 → 유형 묶음 순으로 내려간다.
     구 방식은 「값이 하나라도 있으면 그 층 채택」이라 커버리지 5% 짜리 층이
     뽑혔다 — AELF 는 주제유형 287/287 인데 세부유형 14 행만 있어 273 항목이
     이름 없이 매칭됐다(적중 90%, 이름 단독 22%). 폴백 후 96% / 92%.
     3열 「예측 단위」→「이름 출처」, 코드 접미사는 실제 분할된 행만 부여.

v2 변경 (38단원 실행 전 보완):
  1. unmatched 시트에 중영역·유형 묶음을 함께 실음.
     매칭 컨텍스트는 「중영역 + 유형 묶음 + 유형 이름」인데 v1 회신본에는
     유형 이름만 있어, 규칙 v2 설계 시 어느 층 표지를 잡을지 판단 불가.
     특히 이차곡선처럼 이름 없이 묶음만 있는 단원은 회신이 사실상 공란.
  2. 파이프 분할을 괄호 인식으로 변경. `P(B|A)` 안의 세로줄은 구분자로
     보지 않고 warnings 시트에 보고한다(§4 파이프 충돌).
     원본은 고치지 않는다 — 치환은 사람이 원문비고에 기록하며 한다.
  3. meta에 적용된 단원 팩을 명시. 팩이 없으면(=38단원 대부분) 공통 15개만
     적용됐다는 경고를 남긴다. 적중률만 보고 검증된 7단원과 같은 급으로
     읽히는 것을 막기 위함 (§16 적중률 ≠ 정확도).
  4. detect_layout 의 행 필터가 1열을 그룹ID로 가정하던 것을 헤더 기준으로 수정.
  5. 항목 0건·시트명 불일치 등에서 죽지 않고 이유를 말하고 멈춘다.
  6. 「이름 단독 적중률」과 rule_freq 시트 추가.
     매칭 컨텍스트에 중영역·유형 묶음이 들어가므로, 중영역이 「일차방정식」인
     단원은 C-11 이 전 행에 걸려 유형 정보 없이도 적중률 100%가 나온다.
     이름만으로 몇 %가 걸리는지를 따로 세고, 60%를 넘는 과다 규칙(§6)을
     자동으로 표시한다.
"""
import sys, re, json, os
from openpyxl import load_workbook, Workbook

RULES_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'axis_rules.v3.json')
SHEET_NAME = '문항분류표'
SUMMARY_SHEET = '유형요약'
SCRIPT_VERSION = 'v5'

OPEN_PARENS = '([{（〔'
CLOSE_PARENS = ')]}）〕'


def load_rules(prefix):
    with open(RULES_PATH, encoding='utf-8') as f:
        R = json.load(f)
    rules = list(R['common'])
    packs = []
    for name, pack in R['domain'].items():
        if prefix in pack['applies_to']:
            rules += pack['rules']
            packs.append(name)
    return R['version'], rules, packs


def split_names(raw):
    """세부유형 이름을 파이프로 분리한다(§0. 쉼표 분리 금지).

    괄호 안의 파이프는 구분자가 아니다 — `P(B|A)`, `P(A|B)·P(B)` 등.
    반환: (이름 목록, 괄호 안 파이프가 있었는가)
    """
    parts, buf, depth, guarded = [], [], 0, False
    for ch in raw:
        if ch in OPEN_PARENS:
            depth += 1
        elif ch in CLOSE_PARENS:
            depth = max(0, depth - 1)
        if ch == '|':
            if depth == 0:
                parts.append(''.join(buf))
                buf = []
                continue
            guarded = True
        buf.append(ch)
    parts.append(''.join(buf))
    return [p.strip() for p in parts if p.strip()], guarded


def resolve_name(x):
    """행별로 가장 깊은 유효 이름을 고른다.

    v2 까지는 단원 전체에서 층을 하나 골랐는데, 「값이 하나라도 있으면 채택」
    이라 커버리지 5% 짜리 층이 뽑히고 나머지 행의 이름이 통째로 버려졌다.
    AELF 가 그 경우 — 주제유형 이름 287/287 인데 세부유형 이름 14 행만 있어
    273 항목이 이름 없이 중영역·묶음만으로 매칭됐다(적중 90%, 이름 단독 22%).

    행 단위로 세부유형 → 주제유형 → 유형 묶음 순으로 내려간다.
    주제유형 이름이 그 행의 유형 묶음과 같으면 복제 placeholder 이므로
    출처를 묶음으로 기록한다.
    """
    detail, topic, group = x['세부유형'], x['주제유형'], x['묶음']
    if detail:
        return str(detail), '세부유형'
    if topic and topic != group:
        return str(topic), '주제유형'
    if topic:
        return str(topic), '유형묶음'
    return (str(group), '유형묶음') if group else (None, '없음')


def read_summary_only(wb, cls_rows):
    """유형요약 에만 있고 문항분류표 에 행이 없는 묶음(§16 배지없음)을 방출한다.

    §16 은 배지없는 묶음을 「유형요약에만 등재, 문항분류표 행 없음」으로 정했다.
    그러나 저장 엔진 데이터에는 그 묶음이 그룹으로 존재하므로, xlsx 만 읽으면
    정본 산출물이 엔진보다 커버리지가 낮아진다(45단원 156건).
    묶음 이름 자체가 이름 역할을 하므로 거친 예측은 가능하다.
    """
    if SUMMARY_SHEET not in wb.sheetnames:
        return []
    sy = wb[SUMMARY_SHEET]
    h = {sy.cell(1, c).value: c for c in range(1, sy.max_column + 1)}
    for n in ['그룹ID', '중영역', '유형 묶음']:
        if n not in h:
            return []
    seen = {x['그룹ID'] for x in cls_rows}
    out = []
    for r in range(2, sy.max_row + 1):
        gid = sy.cell(r, h['그룹ID']).value
        if gid is None or gid in seen or str(gid).strip() == '합계':
            continue
        grp = sy.cell(r, h['유형 묶음']).value
        if not grp:
            continue
        out.append({'그룹ID': gid, '코드': gid,   # 주제코드 없음 — 그룹ID로 조인
                    '중영역': sy.cell(r, h['중영역']).value,
                    '묶음': grp, '주제유형': None, '세부유형': None,
                    '요약전용': True})
    return out


def read_rows(ws):
    """열 이름으로 읽는다(§0)."""
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for n in ['그룹ID', '주제코드', '중영역', '유형 묶음']:
        if n not in hdr:
            raise SystemExit(f'열 없음: {n}  (1행 헤더 확인)')
    dcol = hdr.get('세부유형 이름')
    tcol = hdr.get('주제유형 이름')
    out = []
    for r in range(2, ws.max_row + 1):
        gid = ws.cell(r, hdr['그룹ID']).value
        if gid is None:
            continue
        out.append({
            '그룹ID': gid,
            '코드': ws.cell(r, hdr['주제코드']).value,
            '중영역': ws.cell(r, hdr['중영역']).value,
            '묶음': ws.cell(r, hdr['유형 묶음']).value,
            '주제유형': ws.cell(r, tcol).value if tcol else None,
            '세부유형': ws.cell(r, dcol).value if dcol else None,
        })
    return out


def main():
    if len(sys.argv) < 3:
        raise SystemExit(__doc__)
    src, prefix = sys.argv[1], sys.argv[2]
    out = sys.argv[3] if len(sys.argv) > 3 else f'예측표_{prefix}.xlsx'

    if not os.path.exists(RULES_PATH):
        raise SystemExit(f'규칙 파일 없음: {RULES_PATH}\n'
                         '스크립트와 axis_rules.v1.json 은 같은 폴더에 둘 것')
    ver, rules, packs = load_rules(prefix)

    wb_in = load_workbook(src, data_only=True)
    if SHEET_NAME not in wb_in.sheetnames:
        raise SystemExit(f'시트 없음: {SHEET_NAME}  (있는 시트: '
                         f'{", ".join(wb_in.sheetnames)})')
    ws = wb_in[SHEET_NAME]
    src_rows = read_rows(ws)
    summary_only = read_summary_only(wb_in, src_rows)
    src_rows += summary_only

    rows, unmatched, warns = [], [], []
    freq = {}          # 규칙ID → 걸린 항목 수
    name_only_hit = 0  # 중영역·유형 묶음 없이 이름만으로 걸린 항목 수
    layer_count = {}   # 이름 출처 → 항목 수
    pack_gap = []      # 걸렸으나 공통규칙만 걸린 항목
    pack_hit = 0
    for x in src_rows:
        if x.get('요약전용'):
            warns.append([x['그룹ID'], x['코드'], '유형요약 전용 묶음',
                          '문항분류표에 행 없음(§16 배지없음). 묶음 이름으로만 예측. '
                          '주제코드 없으므로 그룹ID로 조인', x['묶음']])
        gid, code, mid, grp = x['그룹ID'], x['코드'], x['중영역'], x['묶음']
        raw, layer = resolve_name(x)

        if raw:
            names, guarded = split_names(raw)
            if guarded:
                warns.append([gid, code, '괄호 안 파이프',
                              '구분자로 보지 않고 이름에 남김. 원본을 ∣(U+2223)로 '
                              '치환할지 §4에 따라 사람이 판단', raw])
            if not names:
                names = [None]
        else:
            names = [None]

        for i, nm in enumerate(names):
            if nm is None or str(nm).strip() == '':
                warns.append([gid, code, '유형 이름 없음',
                              '중영역·유형 묶음만으로 매칭됨 — 변별력 없음', ''])
            layer_count[layer] = layer_count.get(layer, 0) + 1
            # 중영역은 매칭에 쓰지 않는다(v5). 단원 주제어라 변별력이 없다.
            ctx = f'{grp or ""} {nm or ""}'
            axes, hits, whys = [], [], []
            for rule in rules:
                if re.search(rule['pattern'], ctx):
                    hits.append(rule['id'])
                    whys.append(rule['why'])
                    for a in rule['axes']:
                        if a not in axes:
                            axes.append(a)
            for h in hits:
                freq[h] = freq.get(h, 0) + 1
            if nm and any(re.search(rule['pattern'], str(nm))
                          for rule in rules):
                name_only_hit += 1
            key = code if len(names) == 1 else f'{code}.{i + 1:02d}'
            rows.append([gid, key, layer, nm, ', '.join(axes),
                         ', '.join(hits), ' / '.join(dict.fromkeys(whys))])
            if not axes:
                # 규칙 보강용 — 매칭에 쓰인 컨텍스트를 그대로 회신한다
                unmatched.append([gid, key, mid, grp, nm])
            elif any(h.startswith('D-') for h in hits):
                pack_hit += 1
            else:
                # 걸렸지만 단원 고유 함정을 하나도 못 짚은 항목
                pack_gap.append([gid, key, mid, grp, nm, ', '.join(hits)])

    total = len(rows)
    if total == 0:
        raise SystemExit(f'항목 0건. 「{SHEET_NAME}」 시트에 그룹ID가 채워진 '
                         '행이 있는지 확인할 것')
    hit = total - len(unmatched)
    rate = hit / total * 100

    wb = Workbook()
    s = wb.active
    s.title = 'predicted_axes'
    s.append(['그룹ID', '코드', '이름 출처', '유형 이름',
              'predicted_axes', '적용 규칙', '근거'])
    for x in rows:
        s.append(x)
    for col, w in zip('ABCDEFG', [9, 16, 10, 46, 16, 22, 60]):
        s.column_dimensions[col].width = w

    u = wb.create_sheet('unmatched')
    u.append(['그룹ID', '코드', '중영역', '유형 묶음', '유형 이름'])
    for x in unmatched:
        u.append(x)
    for col, w in zip('ABCDE', [9, 16, 22, 34, 52]):
        u.column_dimensions[col].width = w

    p_ = wb.create_sheet('pack_gap')
    p_.append(['그룹ID', '코드', '중영역', '유형 묶음', '유형 이름', '걸린 공통규칙'])
    for x in pack_gap:
        p_.append(x)
    for col, w in zip('ABCDEF', [9, 16, 22, 34, 52, 22]):
        p_.column_dimensions[col].width = w

    w_ = wb.create_sheet('warnings')
    w_.append(['그룹ID', '코드', '종류', '처리', '원문'])
    for x in warns:
        w_.append(x)
    for col, w in zip('ABCDE', [9, 16, 16, 44, 46]):
        w_.column_dimensions[col].width = w

    axis_of = {r['id']: ', '.join(r['axes']) for r in rules}
    over = sorted(((k, v) for k, v in freq.items() if v / total > 0.6),
                  key=lambda x: -x[1])
    f = wb.create_sheet('rule_freq')
    f.append(['규칙ID', '축', '걸린 항목', '비율', '과다(>60%)'])
    for k, v in sorted(freq.items(), key=lambda x: -x[1]):
        f.append([k, axis_of.get(k, ''), v, f'{v / total * 100:.0f}%',
                  '★' if v / total > 0.6 else ''])
    for col, w in zip('ABCDE', [10, 14, 10, 8, 12]):
        f.column_dimensions[col].width = w

    pack_label = ' + '.join(packs) if packs else '없음 (공통 15개만)'
    meta = [('생성 대상', src), ('prefix', prefix), ('규칙 버전', ver),
            ('스크립트 버전', SCRIPT_VERSION),
            ('적용 단원 팩', pack_label),
            ('적용 규칙 수', len(rules)),
            ('이름 출처 분포', ' · '.join(f'{k} {v}' for k, v in sorted(layer_count.items(), key=lambda x: -x[1]))),
            ('총 항목', total), ('적중', hit), ('적중률', f'{rate:.0f}%'),
            ('팩 규칙 걸림', f'{pack_hit} ({pack_hit / total * 100:.0f}%)'),
            ('공통규칙만 걸림', f'{len(pack_gap)} '
                              f'({len(pack_gap) / total * 100:.0f}%)'),
            ('유형요약 전용 묶음', len(summary_only)),
            ('이름 단독 적중', name_only_hit),
            ('이름 단독 적중률', f'{name_only_hit / total * 100:.0f}%'),
            ('과다 규칙(>60%)', ', '.join(f'{k} {v / total * 100:.0f}%'
                                          for k, v in over) or '없음'),
            ('경고 건수', len(warns))]
    if len(pack_gap) / total > 0.5:
        meta.append(('주의', '공통규칙만 걸린 항목이 절반 초과 — 단원 고유 함정을 '
                             '짚지 못하고 있음. 적중률이 높아도 팩이 급함. '
                             'pack_gap 시트가 규칙 보강 재료'))
    if name_only_hit / total < 0.5:
        meta.append(('주의', '이름 단독 적중률 50% 미만 — 적중의 상당수가 '
                             '중영역·유형 묶음 이름에서 나옴. 유형별 변별력이 '
                             '없을 수 있으니 적중률을 그대로 믿지 말 것'))
    if not packs:
        meta.append(('주의', f'{prefix} 전용 단원 팩 없음 — 공통 규칙만 적용. '
                             '적중률이 높아도 정확도는 검증된 7단원과 같은 급이 '
                             '아님 (§8-2 공통만 38%)'))
    meta += [('주의', 'predicted 전용. observed와 같은 필드에 저장 금지'),
             ('주의', '수동 편집 금지. 규칙을 고쳐 재생성할 것'),
             ('주의', '적중률(걸리느냐) ≠ 정확도(맞느냐)'),
             ('주의', '이름은 행별로 세부유형 → 주제유형 → 유형 묶음 순 폴백. '
                      '「단원의 예측 단위」는 더 이상 하나로 정해지지 않음'),
             ('주의', '코드 접미사(.01)는 그 행이 파이프로 분할된 경우에만 붙음. '
                      '조인 키는 주제코드'),
             ('주의', '매칭 컨텍스트는 유형 묶음 + 유형 이름. 중영역은 매칭에 '
                      '쓰지 않음(v5) — 단원 주제어라 전 항목에 같은 축을 붙임')]
    m = wb.create_sheet('meta')
    for k, v in meta:
        m.append([k, v])
    m.column_dimensions['A'].width = 14
    m.column_dimensions['B'].width = 72

    wb.save(out)
    print(f'{prefix}  팩={pack_label}  {total}건  '
          f'적중 {hit} ({rate:.0f}%)  이름단독 {name_only_hit} '
          f'({name_only_hit / total * 100:.0f}%)  경고 {len(warns)}  → {out}')


if __name__ == '__main__':
    main()
